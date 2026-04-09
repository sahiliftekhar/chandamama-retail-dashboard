# ============================================================
#  store/admin.py  —  ChandaMama Retail Intelligence Dashboard
#  With: Date Range Filter + Section Filter + Search
# ============================================================

from django.contrib.admin import DateFieldListFilter
import json
from datetime import timedelta, date
from decimal import Decimal

from django.contrib import admin
from django.db.models import (
    Avg, Count, DecimalField, ExpressionWrapper, F,
    Max, Q, Sum,
)
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.urls import path
from django.utils import timezone
from django.utils.html import format_html

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import AuditLog, Category, Pricing, Product, Sale, Section, Stock, SystemSettings
from .audit_mixin import AuditLogMixin


# ──────────────────────────────────────────────────────────────────
#  CONSTANTS
# ──────────────────────────────────────────────────────────────────

DEAD_STOCK_DAYS = 90


# ──────────────────────────────────────────────────────────────────
#  CUSTOM ADMIN SITE
# ──────────────────────────────────────────────────────────────────

class MyAdminSite(admin.AdminSite):
    site_header = "ChandaMama Retail Dashboard"
    site_title  = "Chandamama ERP"
    index_title = "Operations Centre"
    index_template = "admin/dashboard.html"
    login_template  = "admin/login.html"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path("export-excel/",          self.admin_view(self.export_excel_view),    name="export_excel"),
            path("api/stock-alerts/",      self.admin_view(self.stock_alerts_api),     name="stock_alerts_api"),
            path("api/product-sizes/",     self.admin_view(self.product_sizes_api),    name="product_sizes_api"),
            path("api/dashboard-data/",    self.admin_view(self.dashboard_data_api),   name="dashboard_data_api"),
        ]
        return custom + urls

    # ── INDEX = FULL DASHBOARD ────────────────────────────────────
    def index(self, request, extra_context=None):
        today = timezone.now().date()
        year  = today.year

        # ── Read filters from GET params ──────────────────────────
        date_range   = request.GET.get("range", "year")     # today/week/month/year/all
        section_id   = request.GET.get("section", "all")    # all or section pk
        search_query = request.GET.get("q", "").strip()

        # ── Build date filter ─────────────────────────────────────
        if date_range == "today":
            date_filter = Q(sold_date=today)
            range_label = "Today"
        elif date_range == "week":
            date_filter = Q(sold_date__gte=timezone.now() - timedelta(days=7))
            range_label = "Last 7 Days"
        elif date_range == "month":
            date_filter = Q(sold_date__year=today.year, sold_date__month=today.month)
            range_label = "This Month"
        elif date_range == "all":
            date_filter = Q()
            range_label = "All Time"
        elif date_range == "custom":
            from_date = request.GET.get("from_date")
            to_date   = request.GET.get("to_date")
            try:
                from datetime import datetime
                if from_date and to_date:
                    fd = datetime.strptime(from_date, "%Y-%m-%d").date()
                    td = datetime.strptime(to_date,   "%Y-%m-%d").date()
                    date_filter = Q(sold_date__range=(fd, td))
                    range_label = f"{fd.strftime('%d %b')} → {td.strftime('%d %b %Y')}"
                elif from_date:
                    fd = datetime.strptime(from_date, "%Y-%m-%d").date()
                    date_filter = Q(sold_date=fd)
                    range_label = fd.strftime("%d %b %Y")
                else:
                    date_filter = Q(sold_date=today)
                    range_label = "Today"
            except Exception as e:
                print("DATE ERROR:", e)
                date_filter = Q(sold_date=today)
                range_label = "Today"
        else:  # year (default)
            date_filter = Q(sold_date__year=year)
            range_label = f"FY {year}"

        # ── Build section filter ──────────────────────────────────
        section_filter = Q()
        section_label  = "All Sections"
        all_sections   = Section.objects.all().order_by("name")

        if section_id != "all":
            try:
                sec = Section.objects.get(pk=section_id)
                section_filter = Q(product__category__section=sec)
                section_label  = sec.name
            except Section.DoesNotExist:
                pass

        # ── Combined base queryset ────────────────────────────────
        base_qs = Sale.objects.filter(date_filter & section_filter)

        # ── DEBUG ─────────────────────────────────────────────────
        print("GET PARAMS:", request.GET)
        print("DATE FILTER:", date_filter)
        print("RESULT COUNT:", base_qs.count())

        # ── Search filter (for top products) ─────────────────────
        if search_query:
            search_filter = Q(product__name__icontains=search_query)
        else:
            search_filter = Q()

        # ── KPIs: Today (always today, no filter) ─────────────────
        sales_today = Sale.objects.filter(sold_date=today)
        total_revenue_today = sales_today.aggregate(t=Sum("selling_price"))["t"] or Decimal("0")
        total_profit_today  = sales_today.aggregate(t=Sum("profit"))["t"]        or Decimal("0")
        total_units_today   = sales_today.aggregate(t=Sum("quantity"))["t"]      or 0

        # ── KPIs: Filtered period ─────────────────────────────────
        total_revenue_period = base_qs.aggregate(t=Sum("selling_price"))["t"] or Decimal("0")
        total_profit_period  = base_qs.aggregate(t=Sum("profit"))["t"]        or Decimal("0")
        total_units_period   = base_qs.aggregate(t=Sum("quantity"))["t"]      or 0

        # ── Monthly Revenue (always full year for chart) ──────────
        monthly_base = Sale.objects.filter(sold_date__year=year)
        if section_id != "all":
            monthly_base = monthly_base.filter(section_filter)

        monthly_qs = (
            monthly_base
            .annotate(month=TruncMonth("sold_date"))
            .values("month")
            .annotate(revenue=Sum("selling_price"), profit=Sum("profit"))
            .order_by("month")
        )
        monthly_map_rev    = {e["month"].month: float(e["revenue"] or 0) for e in monthly_qs}
        monthly_map_profit = {e["month"].month: float(e["profit"]  or 0) for e in monthly_qs}
        monthly_labels  = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        monthly_revenue = [monthly_map_rev.get(m, 0)    for m in range(1, 13)]
        monthly_profit  = [monthly_map_profit.get(m, 0) for m in range(1, 13)]

        # ── MoM Growth ────────────────────────────────────────────
        growth_data = [None]
        for i in range(1, 12):
            prev, curr = monthly_revenue[i - 1], monthly_revenue[i]
            if prev and prev > 0:
                growth_data.append(round((curr - prev) / prev * 100, 1))
            else:
                growth_data.append(None)

        # ── Weekly trend (last 7 days, filtered by section) ───────
        weekly_qs = (
            base_qs
            .filter(sold_date__gte=timezone.now() - timedelta(days=7))
            .annotate(day=TruncDate("sold_date"))
            .values("day")
            .annotate(revenue=Sum("selling_price"), profit=Sum("profit"))
            .order_by("day")
        )
        weekly_labels  = [e["day"].strftime("%d %b") for e in weekly_qs]
        weekly_revenue = [float(e["revenue"] or 0) for e in weekly_qs]
        weekly_profit  = [float(e["profit"]  or 0) for e in weekly_qs]

        # ── Category Distribution ─────────────────────────────────
        cat_qs = (
            base_qs
            .values("product__category__name", "product__category__section__name")
            .annotate(revenue=Sum("selling_price"), units=Sum("quantity"))
            .order_by("-revenue")
        )
        total_cat_rev = sum(c["revenue"] for c in cat_qs) or Decimal("1")
        cat_labels = [
            f"{c['product__category__section__name']} › {c['product__category__name']}"
            for c in cat_qs
        ]
        cat_data = [round(float(c["revenue"] / total_cat_rev * 100), 1) for c in cat_qs]

        # ── Top Selling Products (with search) ────────────────────
        top_products = (
            base_qs
            .filter(search_filter)
            .values("product__name", "product__category__name",
                    "product__category__section__name")
            .annotate(units=Sum("quantity"), revenue=Sum("selling_price"), profit=Sum("profit"))
            .order_by("-units")[:10]
        )

        # ── Low Stock Alerts ──────────────────────────────────────
        low_stock_items = []
        stock_qs = Stock.objects.select_related(
            "product", "product__category", "product__category__section"
        ).order_by("quantity")

        if section_id != "all":
            stock_qs = stock_qs.filter(product__category__section_id=section_id)

        for stock in stock_qs:
            threshold = stock.product.low_stock_threshold
            if stock.quantity <= threshold:
                pct = int(stock.quantity / threshold * 100) if threshold else 0
                low_stock_items.append({
                    "product":  stock.product.name,
                    "section":  stock.product.category.section.name
                                if stock.product.category.section else "—",
                    "category": stock.product.category.name,
                    "size":     stock.size,
                    "quantity": stock.quantity,
                    "threshold": threshold,
                    "pct":      pct,
                    "critical": stock.quantity <= max(1, threshold // 3),
                })
        low_stock_count = len(low_stock_items)

        # ── Dead Stock Detection ──────────────────────────────────
        dead_cutoff   = today - timedelta(days=DEAD_STOCK_DAYS)
        sold_recently = Sale.objects.filter(
            sold_date__gte=dead_cutoff
        ).values_list("product_id", flat=True).distinct()

        dead_qs = Stock.objects.filter(quantity__gt=0).exclude(
            product_id__in=sold_recently
        ).select_related(
            "product", "product__category",
            "product__category__section"
        ).annotate(last_sold=Max("product__sale__sold_date"))

        if section_id != "all":
            dead_qs = dead_qs.filter(product__category__section_id=section_id)

        dead_stock_items = []
        for s in dead_qs.order_by(F("last_sold").asc(nulls_first=True))[:20]:
            ls   = s.last_sold.date() if s.last_sold else None
            days = (today - ls).days if ls else None
            try:
                price = float(s.product.pricing.selling_price or 0)
            except Exception:
                price = 0
            dead_stock_items.append({
                "product":    s.product.name,
                "section":    s.product.category.section.name
                              if s.product.category.section else "—",
                "category":   s.product.category.name,
                "size":       s.size,
                "quantity":   s.quantity,
                "last_sold":  ls,
                "days":       days,
                "at_risk":    round(price * s.quantity, 2),
                "truly_dead": days is None or days >= DEAD_STOCK_DAYS,
            })
        dead_stock_count = len(dead_stock_items)

        # ── Product Aging ─────────────────────────────────────────
        def _ago(n): return today - timedelta(days=n)
        prod_qs = Product.objects.all()
        if section_id != "all":
            prod_qs = prod_qs.filter(category__section_id=section_id)

        aging = {
            "fresh":  prod_qs.filter(buy_date__gte=_ago(30)).count(),
            "normal": prod_qs.filter(buy_date__range=[_ago(60), _ago(31)]).count(),
            "warn":   prod_qs.filter(buy_date__range=[_ago(90), _ago(61)]).count(),
            "dead":   prod_qs.filter(buy_date__lt=_ago(90)).count(),
        }

        alert_count = low_stock_count + dead_stock_count

        # ── Profit Margin % by Category ───────────────────────────
        margin_qs = (
            base_qs
            .values("product__category__name", "product__category__section__name")
            .annotate(
                total_revenue=Sum("selling_price"),
                total_profit=Sum("profit"),
            )
            .filter(total_revenue__gt=0)
            .order_by("-total_profit")
        )
        margin_labels  = []
        margin_data    = []
        margin_revenue = []
        margin_profit  = []
        for m in margin_qs:
            if m["total_revenue"] and m["total_revenue"] > 0:
                pct = round(float(m["total_profit"] or 0) / float(m["total_revenue"]) * 100, 1)
                cat_name = f"{m['product__category__section__name']} - {m['product__category__name']}"
                margin_labels.append(cat_name)
                margin_data.append(pct)
                margin_revenue.append(float(m["total_revenue"] or 0))
                margin_profit.append(float(m["total_profit"] or 0))

        context = dict(
            self.each_context(request),
            title="Retail Intelligence Dashboard",
            current_time_ist=timezone.now().strftime("%d %b %Y, %I:%M %p IST"),
            current_date_ist=timezone.now().strftime("%d %B %Y"),
            # KPIs
            total_revenue_today=total_revenue_today,
            total_profit_today=total_profit_today,
            total_units_today=total_units_today,
            total_revenue_period=total_revenue_period,
            total_profit_period=total_profit_period,
            total_units_period=total_units_period,
            # Charts
            monthly_labels=json.dumps(monthly_labels),
            monthly_revenue=json.dumps(monthly_revenue),
            monthly_profit=json.dumps(monthly_profit),
            growth_data=json.dumps(growth_data),
            weekly_labels=json.dumps(weekly_labels),
            weekly_revenue=json.dumps(weekly_revenue),
            weekly_profit=json.dumps(weekly_profit),
            cat_labels=json.dumps(cat_labels),
            cat_data=json.dumps(cat_data),
            # Tables
            top_products=top_products,
            low_stock_items=low_stock_items,
            dead_stock_items=dead_stock_items,
            aging=aging,
            # Counts
            low_stock_count=low_stock_count,
            dead_stock_count=dead_stock_count,
            alert_count=alert_count,
            DEAD_STOCK_DAYS=DEAD_STOCK_DAYS,
            # Filter state
            all_sections=all_sections,
            active_range=date_range,
            active_from_date=request.GET.get("from_date", ""),
            active_to_date=request.GET.get("to_date", ""),
            active_section=section_id,
            range_label=range_label,
            section_label=section_label,
            search_query=search_query,
            # Margin chart
            margin_labels=json.dumps(margin_labels),
            margin_data=json.dumps(margin_data),
            margin_revenue=json.dumps(margin_revenue),
            margin_profit=json.dumps(margin_profit),
        )
        return super().index(request, context)

    # ── EXCEL EXPORT ──────────────────────────────────────────────
    # ── STOCK ALERTS API (for real-time notifications) ────────
    def dashboard_data_api(self, request):
        """AJAX endpoint — returns all dashboard data as JSON for no-reload filtering"""
        from django.http import JsonResponse

        today = timezone.now().date()
        year  = today.year

        date_range   = request.GET.get("range", "year")
        section_id   = request.GET.get("section", "all")
        search_query = request.GET.get("q", "").strip()

        # ── Date filter ──────────────────────────────────────────
        if date_range == "today":
            date_filter = Q(sold_date=today)
            range_label = "Today"
        elif date_range == "week":
            date_filter = Q(sold_date__gte=timezone.now() - timedelta(days=7))
            range_label = "Last 7 Days"
        elif date_range == "month":
            date_filter = Q(sold_date__year=today.year, sold_date__month=today.month)
            range_label = "This Month"
        elif date_range == "all":
            date_filter = Q()
            range_label = "All Time"
        elif date_range == "custom":
            from_date = request.GET.get("from_date")
            to_date   = request.GET.get("to_date")
            try:
                from datetime import datetime as _dt
                if from_date and to_date:
                    fd = _dt.strptime(from_date, "%Y-%m-%d").date()
                    td = _dt.strptime(to_date,   "%Y-%m-%d").date()
                    date_filter = Q(sold_date__range=(fd, td))
                    range_label = f"{fd.strftime('%d %b')} → {td.strftime('%d %b %Y')}"
                elif from_date:
                    fd = _dt.strptime(from_date, "%Y-%m-%d").date()
                    date_filter = Q(sold_date=fd)
                    range_label = fd.strftime("%d %b %Y")
                else:
                    date_filter = Q(sold_date=today)
                    range_label = "Today"
            except Exception:
                date_filter = Q(sold_date=today)
                range_label = "Today"
        else:
            date_filter = Q(sold_date__year=year)
            range_label = f"FY {year}"

        # ── Section filter ───────────────────────────────────────
        section_filter = Q()
        section_label  = "All Sections"
        if section_id != "all":
            try:
                sec = Section.objects.get(pk=section_id)
                section_filter = Q(product__category__section=sec)
                section_label  = sec.name
            except Section.DoesNotExist:
                pass

        base_qs = Sale.objects.filter(date_filter & section_filter)
        search_filter = Q(product__name__icontains=search_query) if search_query else Q()

        # ── KPIs ─────────────────────────────────────────────────
        sales_today = Sale.objects.filter(sold_date=today)
        total_revenue_today  = float(sales_today.aggregate(t=Sum("selling_price"))["t"] or 0)
        total_profit_today   = float(sales_today.aggregate(t=Sum("profit"))["t"] or 0)
        total_units_today    = sales_today.aggregate(t=Sum("quantity"))["t"] or 0
        total_revenue_period = float(base_qs.aggregate(t=Sum("selling_price"))["t"] or 0)
        total_profit_period  = float(base_qs.aggregate(t=Sum("profit"))["t"] or 0)
        total_units_period   = base_qs.aggregate(t=Sum("quantity"))["t"] or 0
        margin_pct = round(total_profit_period / total_revenue_period * 100, 1) if total_revenue_period else 0

        # ── Monthly chart ────────────────────────────────────────
        monthly_base = Sale.objects.filter(sold_date__year=year)
        if section_id != "all":
            monthly_base = monthly_base.filter(section_filter)
        monthly_qs = (monthly_base.annotate(month=TruncMonth("sold_date"))
                      .values("month").annotate(revenue=Sum("selling_price"), profit=Sum("profit"))
                      .order_by("month"))
        monthly_map_rev    = {e["month"].month: float(e["revenue"] or 0) for e in monthly_qs}
        monthly_map_profit = {e["month"].month: float(e["profit"]  or 0) for e in monthly_qs}
        monthly_labels  = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        monthly_revenue = [monthly_map_rev.get(m, 0)    for m in range(1, 13)]
        monthly_profit  = [monthly_map_profit.get(m, 0) for m in range(1, 13)]

        # ── MoM Growth ───────────────────────────────────────────
        growth_data = [None]
        for i in range(1, 12):
            prev, curr = monthly_revenue[i-1], monthly_revenue[i]
            growth_data.append(round((curr-prev)/prev*100, 1) if prev and prev > 0 else None)

        # ── Weekly ───────────────────────────────────────────────
        weekly_qs = (base_qs.filter(sold_date__gte=timezone.now() - timedelta(days=7))
                     .annotate(day=TruncDate("sold_date")).values("day")
                     .annotate(revenue=Sum("selling_price"), profit=Sum("profit")).order_by("day"))
        weekly_labels  = [e["day"].strftime("%d %b") for e in weekly_qs]
        weekly_revenue = [float(e["revenue"] or 0) for e in weekly_qs]
        weekly_profit  = [float(e["profit"]  or 0) for e in weekly_qs]

        # ── Category Distribution ────────────────────────────────
        cat_qs = (base_qs.values("product__category__name", "product__category__section__name")
                  .annotate(revenue=Sum("selling_price")).order_by("-revenue"))
        total_cat_rev = sum(float(c["revenue"] or 0) for c in cat_qs) or 1
        cat_labels = [f"{c['product__category__section__name']} › {c['product__category__name']}" for c in cat_qs]
        cat_data   = [round(float(c["revenue"] or 0) / total_cat_rev * 100, 1) for c in cat_qs]

        # ── Top Products ─────────────────────────────────────────
        top_products = list(base_qs.filter(search_filter)
                            .values("product__name", "product__category__name", "product__category__section__name")
                            .annotate(units=Sum("quantity"), revenue=Sum("selling_price"), profit=Sum("profit"))
                            .order_by("-units")[:10])
        for p in top_products:
            p["revenue"] = float(p["revenue"] or 0)
            p["profit"]  = float(p["profit"]  or 0)

        # ── Low Stock ────────────────────────────────────────────
        low_stock_items = []
        stock_qs = Stock.objects.select_related("product","product__category","product__category__section").order_by("quantity")
        if section_id != "all":
            stock_qs = stock_qs.filter(product__category__section_id=section_id)
        for stock in stock_qs:
            thr = stock.product.low_stock_threshold
            if stock.quantity <= thr:
                pct = int(stock.quantity / thr * 100) if thr else 0
                low_stock_items.append({
                    "product": stock.product.name,
                    "section": stock.product.category.section.name if stock.product.category.section else "—",
                    "category": stock.product.category.name,
                    "size": stock.size, "quantity": stock.quantity,
                    "threshold": thr, "pct": pct,
                    "critical": stock.quantity <= max(1, thr // 3),
                })

        # ── Dead Stock ───────────────────────────────────────────
        dead_cutoff   = today - timedelta(days=DEAD_STOCK_DAYS)
        sold_recently = Sale.objects.filter(sold_date__gte=dead_cutoff).values_list("product_id", flat=True).distinct()
        dead_qs = (Stock.objects.filter(quantity__gt=0).exclude(product_id__in=sold_recently)
                   .select_related("product","product__category","product__category__section")
                   .annotate(last_sold=Max("product__sale__sold_date")))
        if section_id != "all":
            dead_qs = dead_qs.filter(product__category__section_id=section_id)
        dead_stock_items = []
        for s in dead_qs.order_by(F("last_sold").asc(nulls_first=True))[:20]:
            ls   = s.last_sold.date() if s.last_sold else None
            days = (today - ls).days if ls else None
            try: price = float(s.product.pricing.selling_price or 0)
            except: price = 0
            dead_stock_items.append({
                "product": s.product.name,
                "section": s.product.category.section.name if s.product.category.section else "—",
                "category": s.product.category.name,
                "size": s.size, "quantity": s.quantity,
                "last_sold": str(ls) if ls else None, "days": days,
                "at_risk": round(price * s.quantity, 2),
                "truly_dead": days is None or days >= DEAD_STOCK_DAYS,
            })

        # ── Aging ────────────────────────────────────────────────
        def _ago(n): return today - timedelta(days=n)
        prod_qs = Product.objects.all()
        if section_id != "all":
            prod_qs = prod_qs.filter(category__section_id=section_id)
        aging = {
            "fresh":  prod_qs.filter(buy_date__gte=_ago(30)).count(),
            "normal": prod_qs.filter(buy_date__range=[_ago(60), _ago(31)]).count(),
            "warn":   prod_qs.filter(buy_date__range=[_ago(90), _ago(61)]).count(),
            "dead":   prod_qs.filter(buy_date__lt=_ago(90)).count(),
        }

        # ── Margin by Category ───────────────────────────────────
        margin_qs = (base_qs.values("product__category__name","product__category__section__name")
                     .annotate(total_revenue=Sum("selling_price"), total_profit=Sum("profit"))
                     .filter(total_revenue__gt=0).order_by("-total_profit"))
        margin_labels, margin_data, margin_revenue, margin_profit_list = [], [], [], []
        for m in margin_qs:
            if m["total_revenue"] and m["total_revenue"] > 0:
                pct = round(float(m["total_profit"] or 0) / float(m["total_revenue"]) * 100, 1)
                margin_labels.append(f"{m['product__category__section__name']} - {m['product__category__name']}")
                margin_data.append(pct)
                margin_revenue.append(float(m["total_revenue"] or 0))
                margin_profit_list.append(float(m["total_profit"] or 0))

        return JsonResponse({
            "range_label":   range_label,
            "section_label": section_label,
            "active_range":  date_range,
            "active_section": section_id,
            # KPIs
            "total_revenue_today":  total_revenue_today,
            "total_profit_today":   total_profit_today,
            "total_units_today":    total_units_today,
            "total_revenue_period": total_revenue_period,
            "total_profit_period":  total_profit_period,
            "total_units_period":   total_units_period,
            "margin_pct":           margin_pct,
            # Charts
            "monthly_labels":  monthly_labels,
            "monthly_revenue": monthly_revenue,
            "monthly_profit":  monthly_profit,
            "growth_data":     growth_data,
            "weekly_labels":   weekly_labels,
            "weekly_revenue":  weekly_revenue,
            "weekly_profit":   weekly_profit,
            "cat_labels":      cat_labels,
            "cat_data":        cat_data,
            # Tables
            "top_products":     top_products,
            "low_stock_items":  low_stock_items,
            "dead_stock_items": dead_stock_items,
            "low_stock_count":  len(low_stock_items),
            "dead_stock_count": len(dead_stock_items),
            "alert_count":      len(low_stock_items) + len(dead_stock_items),
            "aging":            aging,
            # Margin chart
            "margin_labels":  margin_labels,
            "margin_data":    margin_data,
            "margin_revenue": margin_revenue,
            "margin_profit":  margin_profit_list,
        })

    def product_sizes_api(self, request):
        """Returns available sizes + pricing for a product — used by Sale form JS"""
        from django.http import JsonResponse
        from decimal import Decimal

        product_id = request.GET.get('product_id')
        size_filter = request.GET.get('size')

        if not product_id:
            return JsonResponse({'sizes': []})

        try:
            pricings = Pricing.objects.filter(
                product_id=product_id
            ).order_by('size', 'marked_price')

            if size_filter:
                pricings = pricings.filter(size=size_filter)

            sizes = []
            for p in pricings:
                min_selling = round(float(p.purchase_rate) * 1.20, 2)
                sizes.append({
                    'size':          p.size,
                    'purchase_rate': str(p.purchase_rate),
                    'marked_price':  str(p.marked_price or ''),
                    'min_selling':   str(min_selling),
                    'pricing_id':    p.id,
                })

            return JsonResponse({'sizes': sizes})
        except Exception as e:
            return JsonResponse({'sizes': [], 'error': str(e)})

    def stock_alerts_api(self, request):
        import json as _json
        from datetime import timedelta as _td
        from django.http import JsonResponse
        today = timezone.now().date()
        critical, low, dead = [], [], []
        for stock in Stock.objects.select_related("product","product__category","product__category__section").order_by("quantity"):
            thr = stock.product.low_stock_threshold
            if stock.quantity <= thr:
                item = {"product": stock.product.name, "size": stock.size, "quantity": stock.quantity,
                        "section": stock.product.category.section.name if stock.product.category.section else "",
                        "category": stock.product.category.name}
                if stock.quantity <= max(1, thr // 3): critical.append(item)
                else: low.append(item)
        dead_cutoff = today - timedelta(days=DEAD_STOCK_DAYS)
        sold_ids = Sale.objects.filter(sold_date__gte=dead_cutoff).values_list("product_id", flat=True)
        for s in Stock.objects.filter(quantity__gt=0).exclude(product_id__in=sold_ids).select_related("product","product__category").annotate(last_sold=Max("product__sale__sold_date")).order_by(F("last_sold").asc(nulls_first=True))[:15]:
            ls = s.last_sold.date() if s.last_sold else None
            days = (today - ls).days if ls else None
            try: price = float(s.product.pricing.selling_price or 0)
            except: price = 0
            dead.append({"product": s.product.name, "size": s.size, "quantity": s.quantity, "days": days, "at_risk": round(price * s.quantity, 0)})
        return JsonResponse({"critical": critical, "low": low, "dead": dead})

    def export_excel_view(self, request):
        if not OPENPYXL_AVAILABLE:
            return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

        from openpyxl.styles import Border, Side, GradientFill
        from openpyxl.utils import get_column_letter

        today = timezone.now().date()
        year  = today.year

        # ── Date range from GET params ────────────────────────────
        date_range = request.GET.get("range", "year")
        if date_range == "today":
            sales_filter = Q(sold_date=today)
            period_label = f"Today ({today})"
        elif date_range == "month":
            sales_filter = Q(sold_date__year=year, sold_date__month=today.month)
            period_label = today.strftime("%B %Y")
        elif date_range == "all":
            sales_filter = Q()
            period_label = "All Time"
        else:
            sales_filter = Q(sold_date__year=year)
            period_label = f"FY {year}"

        wb = openpyxl.Workbook()

        # ── Color palette ─────────────────────────────────────────
        HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")
        DARK_FILL  = PatternFill("solid", fgColor="0D1117")
        RED_FILL   = PatternFill("solid", fgColor="3B0010")
        GOLD_FILL  = PatternFill("solid", fgColor="3B2A00")
        GRN_FILL   = PatternFill("solid", fgColor="003B15")
        TOTAL_FILL = PatternFill("solid", fgColor="1A2744")
        ALT_FILL   = PatternFill("solid", fgColor="F0F4F8")
        ALT_FILL2  = PatternFill("solid", fgColor="E8F4FD")

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def make_header(ws, headers, subtitle=""):
            # Title row
            ws.row_dimensions[1].height = 35
            title_cell = ws.cell(row=1, column=1, value=f"ChandaMama Retail Dashboard — {ws.title}")
            title_cell.font      = Font(bold=True, color="00D4FF", size=14, name="Calibri")
            title_cell.fill      = DARK_FILL
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

            # Subtitle row
            ws.row_dimensions[2].height = 20
            sub_cell = ws.cell(row=2, column=1, value=f"Period: {period_label}  |  Generated: {today}")
            sub_cell.font      = Font(color="7A9BB5", size=9, italic=True, name="Calibri")
            sub_cell.fill      = DARK_FILL
            sub_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

            # Empty row
            ws.row_dimensions[3].height = 8
            for col in range(1, len(headers)+1):
                ws.cell(row=3, column=col).fill = DARK_FILL

            # Header row
            ws.row_dimensions[4].height = 28
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
                cell.fill      = HDR_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = border

        def add_total_row(ws, start_col, end_col, data_start=5):
            last_row  = ws.max_row
            total_row = last_row + 1
            ws.row_dimensions[total_row].height = 22
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="00D4FF", size=10, name="Calibri")
            ws.cell(row=total_row, column=1).fill = TOTAL_FILL
            ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
            for col in range(start_col, end_col+1):
                cell = ws.cell(row=total_row, column=col)
                col_letter = get_column_letter(col)
                cell.value  = f"=SUM({col_letter}{data_start}:{col_letter}{last_row})"
                cell.font   = Font(bold=True, color="00FF88", size=10, name="Calibri")
                cell.fill   = TOTAL_FILL
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

        def auto_width(ws):
            for col in ws.columns:
                try:
                    col_letter = col[0].column_letter
                    w = max((len(str(c.value or "")) for c in col if hasattr(c, 'value')), default=10)
                    ws.column_dimensions[col_letter].width = min(w + 6, 45)
                except AttributeError:
                    pass

        def add_row_colors(ws, data_start=5, profit_col=None, alt=True):
            for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, max_row=ws.max_row), 1):
                for cell in row:
                    cell.border = border
                    cell.font   = cell.font.copy(name="Calibri") if cell.font else Font(name="Calibri")
                    if alt and row_idx % 2 == 0:
                        if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                            cell.fill = ALT_FILL
                if profit_col:
                    profit_cell = row[profit_col - 1]
                    try:
                        val = float(profit_cell.value or 0)
                        if val > 0:
                            profit_cell.font = Font(bold=True, color="00AA55", name="Calibri")
                        elif val < 0:
                            profit_cell.font = Font(bold=True, color="FF3B5C", name="Calibri")
                    except (TypeError, ValueError):
                        pass

        # ── Sheet 1: Summary Dashboard ────────────────────────────
        ws0 = wb.active
        ws0.title = "Summary"
        headers0 = ["Metric", "Value", "Notes"]
        make_header(ws0, headers0)

        total_rev   = Sale.objects.filter(sales_filter).aggregate(t=Sum("selling_price"))["t"] or 0
        total_prof  = Sale.objects.filter(sales_filter).aggregate(t=Sum("profit"))["t"] or 0
        total_units = Sale.objects.filter(sales_filter).aggregate(t=Sum("quantity"))["t"] or 0
        total_orders= Sale.objects.filter(sales_filter).count()
        today_rev   = Sale.objects.filter(sold_date=today).aggregate(t=Sum("selling_price"))["t"] or 0
        today_prof  = Sale.objects.filter(sold_date=today).aggregate(t=Sum("profit"))["t"] or 0
        low_count   = sum(1 for s in Stock.objects.select_related("product") if s.quantity <= s.product.low_stock_threshold)
        margin_pct  = round(float(total_prof)/float(total_rev)*100,1) if total_rev else 0
        avg_order   = round(float(total_rev)/total_orders,2) if total_orders else 0

        summary_rows = [
            ("📅 Report Date",       str(today),              "Generated automatically"),
            ("📆 Period",            period_label,            "Selected date range"),
            ("",                     "",                      ""),
            ("💰 Total Revenue",     float(total_rev),        f"₹{total_rev:,.2f}"),
            ("📈 Total Profit",      float(total_prof),       f"₹{total_prof:,.2f}"),
            ("📊 Profit Margin",     f"{margin_pct}%",        "Net margin percentage"),
            ("🛒 Units Sold",        total_units,             "Total items sold"),
            ("📋 Transactions",      total_orders,            "Total sale records"),
            ("💵 Avg Order Value",   float(avg_order),        f"₹{avg_order:,.2f} per transaction"),
            ("",                     "",                      ""),
            ("📅 Today Revenue",     float(today_rev),        str(today)),
            ("📅 Today Profit",      float(today_prof),       str(today)),
            ("",                     "",                      ""),
            ("⚠️ Low Stock SKUs",   low_count,               "Items below threshold"),
        ]

        # Section-wise breakdown
        ws0.cell(row=len(summary_rows)+6, column=1, value="SECTION BREAKDOWN").font = Font(bold=True, color="00D4FF", size=11, name="Calibri")
        sec_data = Sale.objects.filter(sales_filter).values("product__category__section__name").annotate(rev=Sum("selling_price"), profit=Sum("profit"), units=Sum("quantity")).order_by("-rev")
        for i, row_data in enumerate(summary_rows, 5):
            for col, val in enumerate(row_data, 1):
                cell = ws0.cell(row=i, column=col, value=val)
                cell.font   = Font(name="Calibri", size=10)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if col == 1:
                    cell.font = Font(bold=True, name="Calibri", size=10, color="1E3A5F")
                if col == 2 and isinstance(val, float) and val > 0:
                    cell.number_format = "#,##0.00"
                if i % 2 == 0:
                    cell.fill = ALT_FILL

        auto_width(ws0)

        # ── Sheet 2: All Sales ────────────────────────────────────
        ws1 = wb.create_sheet("All Sales")
        hdrs1 = ["#","Product","Section","Category","Size","Qty","Selling Price (₹)","Purchase Rate (₹)","Discount (₹)","Profit (₹)","Margin %","Date"]
        make_header(ws1, hdrs1)
        sales_qs = Sale.objects.select_related("product","product__category","product__category__section").filter(sales_filter).order_by("-sold_date")
        for i, s in enumerate(sales_qs, 1):
            margin = round(float(s.profit or 0)/float(s.purchase_rate_snapshot * s.quantity)*100,1) if s.purchase_rate_snapshot and s.quantity else 0
            ws1.append([
                i, s.product.name,
                s.product.category.section.name if s.product.category.section else "",
                s.product.category.name, s.size, s.quantity,
                float(s.selling_price), float(s.purchase_rate_snapshot or 0),
                float(s.discount), float(s.profit or 0), f"{margin}%",
                s.sold_date.strftime("%Y-%m-%d %H:%M"),
            ])
        add_total_row(ws1, 6, 10)
        add_row_colors(ws1, profit_col=10)
        auto_width(ws1)

        # ── Sheet 3: Monthly Revenue ──────────────────────────────
        ws2 = wb.create_sheet("Monthly Revenue")
        hdrs2 = ["Month","Revenue (₹)","Profit (₹)","Margin %","Units Sold","Transactions","Avg Sale (₹)"]
        make_header(ws2, hdrs2)
        mnames = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        monthly_data = (
            Sale.objects.filter(sold_date__year=year)
            .annotate(month=TruncMonth("sold_date"))
            .values("month")
            .annotate(rev=Sum("selling_price"), profit=Sum("profit"),
                      units=Sum("quantity"), txns=Count("id"), avg=Avg("selling_price"))
            .order_by("month")
        )
        for m in monthly_data:
            rev = float(m["rev"] or 0)
            prof = float(m["profit"] or 0)
            marg = round(prof/rev*100,1) if rev else 0
            ws2.append([mnames[m["month"].month-1], rev, prof, f"{marg}%", m["units"], m["txns"], float(m["avg"] or 0)])
        add_total_row(ws2, 2, 3)
        add_row_colors(ws2, profit_col=3)
        auto_width(ws2)

        # ── Sheet 4: Top Products ─────────────────────────────────
        ws3 = wb.create_sheet("Top Products")
        hdrs3 = ["Rank","Product","Section","Category","Units Sold","Revenue (₹)","Profit (₹)","Margin %"]
        make_header(ws3, hdrs3)
        top_prods = (
            Sale.objects.filter(sales_filter)
            .values("product__name","product__category__name","product__category__section__name")
            .annotate(units=Sum("quantity"), rev=Sum("selling_price"), profit=Sum("profit"))
            .order_by("-units")[:50]
        )
        for i, p in enumerate(top_prods, 1):
            rev  = float(p["rev"] or 0)
            prof = float(p["profit"] or 0)
            marg = round(prof/rev*100,1) if rev else 0
            row  = [i, p["product__name"], p["product__category__section__name"] or "",
                    p["product__category__name"], p["units"], rev, prof, f"{marg}%"]
            ws3.append(row)
            for cell in ws3[ws3.max_row]:
                cell.border = border
                cell.font   = Font(name="Calibri", size=10)
                if i == 1:
                    cell.fill = PatternFill("solid", fgColor="1A3A1A")
                    cell.font = Font(bold=True, color="00FF88", name="Calibri", size=10)
                elif i == 2:
                    cell.fill = PatternFill("solid", fgColor="1A2A3A")
                    cell.font = Font(bold=True, color="00D4FF", name="Calibri", size=10)
                elif i == 3:
                    cell.fill = PatternFill("solid", fgColor="3A2A1A")
                    cell.font = Font(bold=True, color="FFB347", name="Calibri", size=10)
                elif i % 2 == 0:
                    cell.fill = ALT_FILL
        auto_width(ws3)

        # ── Sheet 5: Category Analysis ────────────────────────────
        ws_cat = wb.create_sheet("Category Analysis")
        hdrs_cat = ["Section","Category","Units Sold","Revenue (₹)","Profit (₹)","Margin %","% of Total Revenue"]
        make_header(ws_cat, hdrs_cat)
        cat_data = (
            Sale.objects.filter(sales_filter)
            .values("product__category__section__name","product__category__name")
            .annotate(units=Sum("quantity"), rev=Sum("selling_price"), profit=Sum("profit"))
            .order_by("-rev")
        )
        total_cat_rev = float(total_rev) or 1
        for c in cat_data:
            rev  = float(c["rev"] or 0)
            prof = float(c["profit"] or 0)
            marg = round(prof/rev*100,1) if rev else 0
            pct  = round(rev/total_cat_rev*100,1)
            ws_cat.append([c["product__category__section__name"] or "", c["product__category__name"],
                           c["units"], rev, prof, f"{marg}%", f"{pct}%"])
        add_total_row(ws_cat, 3, 5)
        add_row_colors(ws_cat, profit_col=5)
        auto_width(ws_cat)

        # ── Sheet 6: Section-wise Revenue ─────────────────────────
        ws_sec = wb.create_sheet("Section Analysis")
        hdrs_sec = ["Section","Units Sold","Revenue (₹)","Profit (₹)","Margin %","Transactions","Avg Sale (₹)"]
        make_header(ws_sec, hdrs_sec)
        sec_data = (
            Sale.objects.filter(sales_filter)
            .values("product__category__section__name")
            .annotate(units=Sum("quantity"), rev=Sum("selling_price"),
                      profit=Sum("profit"), txns=Count("id"), avg=Avg("selling_price"))
            .order_by("-rev")
        )
        for s in sec_data:
            rev  = float(s["rev"] or 0)
            prof = float(s["profit"] or 0)
            marg = round(prof/rev*100,1) if rev else 0
            ws_sec.append([s["product__category__section__name"] or "Unknown",
                           s["units"], rev, prof, f"{marg}%", s["txns"], float(s["avg"] or 0)])
        add_total_row(ws_sec, 2, 4)
        add_row_colors(ws_sec, profit_col=4)
        auto_width(ws_sec)

        # ── Sheet 7: Low Stock Alerts ─────────────────────────────
        ws4 = wb.create_sheet("Low Stock Alerts")
        hdrs4 = ["Product","Section","Category","Size","Current Qty","Threshold","Shortage","Status","Urgency"]
        make_header(ws4, hdrs4)
        for stk in Stock.objects.select_related("product","product__category","product__category__section").order_by("quantity"):
            thr = stk.product.low_stock_threshold
            if stk.quantity <= thr:
                shortage = thr - stk.quantity
                status   = "CRITICAL" if stk.quantity <= max(1, thr//3) else "LOW"
                urgency  = "🔴 Order Now" if status == "CRITICAL" else "🟡 Order Soon"
                ws4.append([
                    stk.product.name,
                    stk.product.category.section.name if stk.product.category.section else "",
                    stk.product.category.name,
                    stk.size, stk.quantity, thr, shortage, status, urgency,
                ])
                fill = RED_FILL if status == "CRITICAL" else GOLD_FILL
                for cell in ws4[ws4.max_row]:
                    cell.fill   = fill
                    cell.border = border
                    cell.font   = Font(name="Calibri", size=10,
                                       color="FF6B6B" if status == "CRITICAL" else "FFD700")
        auto_width(ws4)

        # ── Sheet 8: Dead Stock ───────────────────────────────────
        ws5 = wb.create_sheet("Dead Stock")
        hdrs5 = ["Product","Section","Category","Size","Qty","Last Sold","Days Since Sale","At-Risk Value (₹)","Action"]
        make_header(ws5, hdrs5)
        dead_cutoff = today - timedelta(days=DEAD_STOCK_DAYS)
        sold_ids    = Sale.objects.filter(sold_date__gte=dead_cutoff).values_list("product_id", flat=True)
        for stk in (
            Stock.objects.filter(quantity__gt=0).exclude(product_id__in=sold_ids)
            .select_related("product","product__category","product__category__section")
            .annotate(last_sold=Max("product__sale__sold_date"))
            .order_by(F("last_sold").asc(nulls_first=True))
        ):
            ls   = stk.last_sold.date() if stk.last_sold else None
            days = (today - ls).days if ls else "Never"
            try: price = float(stk.product.pricing.selling_price or 0)
            except: price = 0
            action = "Consider Discount" if isinstance(days, int) and days < 120 else "Clear Stock"
            ws5.append([
                stk.product.name,
                stk.product.category.section.name if stk.product.category.section else "",
                stk.product.category.name,
                stk.size, stk.quantity,
                str(ls) if ls else "Never", days,
                round(price * stk.quantity, 2), action,
            ])
            for cell in ws5[ws5.max_row]:
                cell.fill   = GOLD_FILL
                cell.border = border
                cell.font   = Font(name="Calibri", size=10, color="FFD700")
        auto_width(ws5)

        # ── Freeze panes on all sheets ────────────────────────────
        for ws in wb.worksheets:
            ws.freeze_panes = "A5"

        filename = f"chandamama_report_{date_range}_{today}.xlsx"
        resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp


my_admin = MyAdminSite(name="my_admin")


# ──────────────────────────────────────────────────────────────────
#  INLINES
# ──────────────────────────────────────────────────────────────────

class StockInline(admin.TabularInline):
    model   = Stock
    extra   = 3
    min_num = 0
    fields  = ("size", "quantity")
    verbose_name        = "Size & Stock"
    verbose_name_plural = "Sizes & Stock"

    def get_extra(self, request, obj=None, **kwargs):
        # Don't show extra empty rows if stock already exists
        if obj and obj.stock_set.exists():
            return 0
        return 1


class PricingInline(admin.TabularInline):
    model   = Pricing
    extra   = 1
    min_num = 0
    # ✅ Fix 1: size added so each size can have its own price
    # ✅ Fix 3: selling_price removed — it's entered only during EOD sale
    fields  = ("size", "purchase_rate", "marked_price")
    verbose_name        = "Size Pricing"
    verbose_name_plural = "Size-based Pricing (add one row per size)"

    def get_extra(self, request, obj=None, **kwargs):
        # Don't show extra empty rows if pricing already exists
        if obj and obj.pricings.exists():
            return 0
        return 1


# ──────────────────────────────────────────────────────────────────
#  MODEL ADMINS
# ──────────────────────────────────────────────────────────────────

@admin.register(Section, site=my_admin)
class SectionAdmin(AuditLogMixin, admin.ModelAdmin):
    list_display  = ("name", "category_count")
    search_fields = ("name",)

    @admin.display(description="Categories")
    def category_count(self, obj):
        return obj.category_set.count()


@admin.register(Category, site=my_admin)
class CategoryAdmin(AuditLogMixin, admin.ModelAdmin):
    list_display  = ("name", "section", "product_count", "total_revenue")
    list_filter   = ("section",)
    search_fields = ("name",)

    @admin.display(description="Products")
    def product_count(self, obj):
        return obj.product_set.count()

    @admin.display(description="Revenue")
    def total_revenue(self, obj):
        val = Sale.objects.filter(
            product__category=obj
        ).aggregate(t=Sum("selling_price"))["t"] or 0
        val_str = "{:,.2f}".format(float(val))
        return format_html('<span style="color:#00d4ff">₹{}</span>', val_str)


@admin.register(Product, site=my_admin)
class ProductAdmin(AuditLogMixin, admin.ModelAdmin):
    inlines       = [PricingInline, StockInline]
    list_display  = ("name", "category_section", "buy_date",
                     "low_stock_threshold", "total_stock", "stock_status_badge")
    list_filter   = ("category__section", "category")
    search_fields = ("name",)
    readonly_fields = ()

    @admin.display(description="Section › Category")
    def category_section(self, obj):
        sec = obj.category.section.name if obj.category.section else "—"
        return f"{sec} › {obj.category.name}"

    @admin.display(description="Total Stock")
    def total_stock(self, obj):
        total = obj.stock_set.aggregate(t=Sum("quantity"))["t"] or 0
        return format_html("<strong>{}</strong>", total)

    @admin.display(description="Status")
    def stock_status_badge(self, obj):
        total = obj.stock_set.aggregate(t=Sum("quantity"))["t"] or 0
        thr   = obj.low_stock_threshold
        if total == 0:
            label, color, bg = "OUT OF STOCK", "#ff3b5c", "rgba(255,59,92,0.1)"
        elif total <= max(1, thr // 3):
            label, color, bg = "CRITICAL",     "#ff3b5c", "rgba(255,59,92,0.1)"
        elif total <= thr:
            label, color, bg = "LOW STOCK",    "#ff6b35", "rgba(255,107,53,0.1)"
        else:
            label, color, bg = "OK",           "#00ff88", "rgba(0,255,136,0.1)"
        return format_html(
            '<span style="padding:3px 10px;border-radius:20px;font-size:10px;'
            'font-weight:700;color:{};background:{};border:1px solid {}">{}</span>',
            color, bg, color, label,
        )

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            "category", "category__section"
        ).prefetch_related("stock_set")

    def save_formset(self, request, form, formset, change):
        """✅ Fix 3: Expand size ranges when saving Stock and Pricing inlines"""
        from .models import expand_size_range
        instances = formset.save(commit=False)

        for instance in instances:
            if isinstance(instance, Stock):
                sizes = expand_size_range(instance.size)
                if len(sizes) > 1:
                    # It's a range — create individual stock entries
                    for sz in sizes:
                        existing = Stock.objects.filter(
                            product=instance.product, size=sz
                        ).first()
                        if existing:
                            existing.quantity += instance.quantity
                            existing.save()
                        else:
                            Stock.objects.create(
                                product=instance.product,
                                size=sz,
                                quantity=instance.quantity
                            )
                    # Don't save the range entry itself
                    continue
                else:
                    # Normal size — save as usual
                    existing = Stock.objects.filter(
                        product=instance.product, size=instance.size
                    ).exclude(pk=instance.pk if instance.pk else None).first()
                    if existing:
                        existing.quantity += instance.quantity
                        existing.save()
                        continue

            elif isinstance(instance, Pricing):
                sizes = expand_size_range(instance.size)
                if len(sizes) > 1:
                    # Range pricing — create one pricing per size
                    for sz in sizes:
                        Pricing.objects.create(
                            product=instance.product,
                            size=sz,
                            purchase_rate=instance.purchase_rate,
                            marked_price=instance.marked_price
                        )
                    continue

            instance.save()

        formset.save_m2m()

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        return form

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        # ✅ Fix 2: On POST, check for duplicate and redirect to existing
        if request.method == 'POST' and not object_id:
            name     = request.POST.get('name', '').strip()
            category = request.POST.get('category')
            if name and category:
                existing = Product.objects.filter(
                    name=name, category_id=category
                ).first()
                if existing:
                    from django.contrib import messages
                    from django.http import HttpResponseRedirect
                    from django.urls import reverse
                    messages.warning(
                        request,
                        f'⚠️ "{name}" already exists in this category. '
                        f'You have been redirected to the existing product — '
                        f'add new sizes/prices here.'
                    )
                    url = reverse('my_admin:store_product_change',
                                  args=[existing.id])
                    return HttpResponseRedirect(url)
        return super().changeform_view(
            request, object_id, form_url, extra_context
        )


@admin.register(Pricing, site=my_admin)
class PricingAdmin(AuditLogMixin, admin.ModelAdmin):
    list_display  = ("product", "size", "purchase_rate", "marked_price", "margin_display")
    search_fields = ("product__name",)

    @admin.display(description="Margin %")
    def margin_display(self, obj):
        try:
            pct   = obj.margin_percent()
            color = "#00ff88" if pct >= 30 else "#fbbf24" if pct >= 20 else "#ff3b5c"
            return format_html(
                '<span style="color:{};font-weight:700">{}%</span>', color, pct
            )
        except Exception:
            return "—"


@admin.register(Stock, site=my_admin)
class StockAdmin(AuditLogMixin, admin.ModelAdmin):
    list_display  = ("product", "section_name", "size", "quantity", "stock_bar")
    list_filter   = ("product__category__section", "product__category")
    search_fields = ("product__name", "size")
    ordering      = ("quantity",)

    @admin.display(description="Section")
    def section_name(self, obj):
        return obj.product.category.section.name if obj.product.category.section else "—"

    @admin.display(description="Stock Level")
    def stock_bar(self, obj):
        thr   = obj.product.low_stock_threshold
        pct   = min(int(obj.quantity / thr * 100), 100) if thr else 100
        color = "#ff3b5c" if obj.quantity <= max(1, thr//3) else \
                "#ff6b35" if obj.quantity <= thr else "#00ff88"
        return format_html(
            '<div style="display:flex;align-items:center;gap:8px">'
            '<div style="width:100px;height:6px;background:#1e2d3d;border-radius:3px">'
            '<div style="width:{}%;height:100%;background:{};border-radius:3px"></div>'
            '</div><span style="color:{};font-size:11px;font-weight:700">{}%</span>'
            '</div>',
            pct, color, color, pct,
        )

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            "product", "product__category", "product__category__section"
        )


@admin.register(Sale, site=my_admin)
class SaleAdmin(AuditLogMixin, admin.ModelAdmin):
    list_display = ("product", "size", "quantity", "selling_price_display",
                    "profit_display", "margin_live", "payment_badge", "sold_date")
    list_filter = (
        ("sold_date", DateFieldListFilter),
        "product__category__section",
        "product__category",
    )

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['cart_url'] = '/admin/store/sale/add-cart/'
        return super().changelist_view(request, extra_context=extra_context)

    readonly_fields = ("profit", "purchase_rate_snapshot", "marked_price_snapshot")
    date_hierarchy  = "sold_date"

    fields = (
        ("product", "size"),
        ("quantity", "selling_price", "discount"),
        ("payment_mode",),
        ("customer_name", "customer_phone"),
        ("sold_date",),
        ("pricing_id",),
        ("profit", "purchase_rate_snapshot", "marked_price_snapshot"),
    )

    @admin.display(description="Payment")
    def payment_badge(self, obj):
        colors = {
            'cash':     '#00ff88',
            'phonepay': '#007bff',
            'due':      '#ff3b5c',
        }
        labels = {
            'cash':     '💵 Cash',
            'phonepay': '📱 PhonePe',
            'due':      '⏳ Due',
        }
        color = colors.get(obj.payment_mode, '#aaa')
        label = labels.get(obj.payment_mode, obj.payment_mode)
        return format_html(
            '<span style="padding:3px 10px;border-radius:20px;'
            'font-size:11px;font-weight:600;background:{};color:#000">{}</span>',
            color, label
        )

    # ✅ Fix 2: Dynamic size choices based on selected product
    class Media:
        js = ('admin/js/sale_size_selector.js',)

    @admin.display(description="Selling Price", ordering="selling_price")
    def selling_price_display(self, obj):
        return format_html(
            '<span style="color:#00d4ff;font-weight:600">₹{}</span>', obj.selling_price
        )

    @admin.display(description="Profit", ordering="profit")
    def profit_display(self, obj):
        if obj.profit is None: return "—"
        color = "#00ff88" if obj.profit > 0 else "#ff3b5c"
        return format_html('<strong style="color:{}">₹{}</strong>', color, obj.profit)

    @admin.display(description="Margin %")
    def margin_live(self, obj):
        try:
            pct   = round(float(obj.profit or 0) /
                          float(obj.purchase_rate_snapshot * obj.quantity) * 100, 1)
            color = "#00ff88" if pct >= 30 else "#fbbf24" if pct >= 20 else "#ff3b5c"
            return format_html(
                '<span style="color:{};font-weight:700">{}%</span>', color, pct
            )
        except Exception:
            return "—"

    @admin.display(description="Payment")
    def payment_badge(self, obj):
        colors = {
            'cash':     '#00ff88',
            'phonepay': '#007bff',
            'due':      '#ff3b5c',
        }
        labels = {
            'cash':     '💵 Cash',
            'phonepay': '📱 PhonePe',
            'due':      '⏳ Due',
        }
        color = colors.get(obj.payment_mode, '#aaa')
        label = labels.get(obj.payment_mode, obj.payment_mode)
        return format_html(
            '<span style="padding:3px 10px;border-radius:20px;'
            'font-size:11px;font-weight:600;background:{};color:#000">{}</span>',
            color, label
        )

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if 'pricing_id' in form.base_fields:
            form.base_fields['pricing_id'].widget.attrs['style'] = 'display:none'
            form.base_fields['pricing_id'].label = ''
            form.base_fields['pricing_id'].required = False
        return form

    def get_urls(self):
        from django.urls import path
        custom = [
            path("add-cart/",      self.admin_site.admin_view(self.cart_view),     name="store_sale_cart"),
            path("add-cart/save/", self.admin_site.admin_view(self.cart_save_api), name="store_sale_cart_save"),
        ]
        return custom + super().get_urls()
    
    def cart_view(self, request):
        products = Product.objects.all().order_by("name")
        import json as _json
        products_json = _json.dumps([{"id": p.id, "name": p.name} for p in products])
        context = dict(
            self.admin_site.each_context(request),
            title="Add Sale — Cart",
            products=products,
            products_json=products_json,
            opts=self.model._meta,
        )
        return TemplateResponse(request, "admin/store/sale/cart.html", context)

    def cart_save_api(self, request):
        from django.http import JsonResponse
        import json as _json

        if request.method != "POST":
            return JsonResponse({"error": "POST required"}, status=405)

        try:
            data = _json.loads(request.body)
        except Exception:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        items          = data.get("items", [])
        payment_mode   = data.get("payment_mode", "cash")
        customer_name  = data.get("customer_name", "").strip()
        customer_phone = data.get("customer_phone", "").strip()
        sold_date_str  = data.get("sold_date", "")

        if not items:
            return JsonResponse({"error": "No items in cart"}, status=400)

        if payment_mode == "due" and (not customer_name or not customer_phone):
            return JsonResponse({"error": "Customer name and phone required for Due sales"}, status=400)

        saved  = []
        errors = []

        for item in items:
            try:
                from datetime import date as _date
                sale = Sale(
                    product_id    = item.get("product_id"),
                    size          = item.get("size"),
                    quantity      = int(item.get("quantity", 1)),
                    selling_price = Decimal(str(item.get("selling_price", 0))),
                    discount      = Decimal(str(item.get("discount", 0))),
                    payment_mode  = payment_mode,
                    customer_name  = customer_name if payment_mode == "due" else "",
                    customer_phone = customer_phone if payment_mode == "due" else "",
                    pricing_id    = item.get("pricing_id"),
                )
                if sold_date_str:
                    sale.sold_date = _date.fromisoformat(sold_date_str)
                sale.full_clean()
                sale.save()
                saved.append(str(sale))
            except Exception as e:
                errors.append({"item": item.get("product_id"), "error": str(e)})

        if errors and not saved:
            return JsonResponse({"error": errors[0]["error"], "all_errors": errors}, status=400)

        return JsonResponse({
            "saved": len(saved),
            "errors": errors,
            "message": f"{len(saved)} sale(s) saved successfully!"
        })

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            "product", "product__category", "product__category__section"
        )


@admin.register(SystemSettings, site=my_admin)
class SystemSettingsAdmin(admin.ModelAdmin):
    list_display = ("__str__", "show_purchase_rate")


# ──────────────────────────────────────────────────────────────────
#  AUDIT LOG ADMIN
# ──────────────────────────────────────────────────────────────────

@admin.register(AuditLog, site=my_admin)
class AuditLogAdmin(admin.ModelAdmin):
    list_display  = ("timestamp_display", "user", "action_badge",
                     "model_name", "object_repr_short", "changes_short", "ip_address")
    list_filter   = ("action", "model_name", "user", "timestamp")
    search_fields = ("object_repr", "changes", "user__username", "model_name")
    readonly_fields = ("user", "action", "model_name", "object_id",
                       "object_repr", "changes", "ip_address", "timestamp")
    date_hierarchy = "timestamp"
    ordering      = ("-timestamp",)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser

    @admin.display(description="Time", ordering="timestamp")
    def timestamp_display(self, obj):
        return format_html(
            '<span style="font-family:\'JetBrains Mono\',monospace;font-size:11px;color:#64748b">{}</span>',
            obj.timestamp.strftime("%d %b %Y %H:%M")
        )

    @admin.display(description="Action")
    def action_badge(self, obj):
        config = {
            "CREATE": ("#00ff88", "rgba(0,255,136,0.1)", "✚ CREATE"),
            "UPDATE": ("#00d4ff", "rgba(0,212,255,0.1)", "✎ UPDATE"),
            "DELETE": ("#ff3b5c", "rgba(255,59,92,0.1)", "✖ DELETE"),
        }
        color, bg, label = config.get(obj.action, ("#94a3b8", "transparent", obj.action))
        return format_html(
            '<span style="padding:3px 9px;border-radius:20px;font-size:10px;'
            'font-weight:700;color:{};background:{};border:1px solid {}">{}</span>',
            color, bg, color, label,
        )

    @admin.display(description="Object")
    def object_repr_short(self, obj):
        text = obj.object_repr[:60] + "..." if len(obj.object_repr) > 60 else obj.object_repr
        return format_html('<span style="font-weight:600">{}</span>', text)

    @admin.display(description="Changes")
    def changes_short(self, obj):
        if not obj.changes:
            return format_html('<span style="color:#94a3b8">—</span>')
        text = obj.changes[:80] + "..." if len(obj.changes) > 80 else obj.changes
        return format_html(
            '<span style="font-family:\'JetBrains Mono\',monospace;font-size:10px;color:#64748b">{}</span>',
            text
        )